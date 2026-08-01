"""
นำเข้าคะแนนจากไฟล์ Excel (เทมเพลต score_template.xlsx) เข้าสู่ฐานข้อมูล SQLite

วิธีใช้:
    python import_scores.py path/to/scores.xlsx [--sheet ชื่อชีท] [--replace]

--replace  ล้างข้อมูลนักเรียน/คะแนนเดิมทั้งหมดก่อนนำเข้าใหม่ (ค่าเริ่มต้นคือเพิ่ม/อัปเดตทับ)
"""
import argparse
import sys
import openpyxl

from db import get_conn, init_db

INFO_COLS = ["ชั้น", "เลขที่", "ชื่อ-สกุล"]


def load_sheet(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    elif "คะแนน" in wb.sheetnames:
        ws = wb["คะแนน"]
    else:
        ws = wb[wb.sheetnames[0]]
    return ws


def import_scores(path, sheet_name=None, replace=False):
    ws = load_sheet(path, sheet_name)

    header = [c.value for c in ws[1]]
    if header[:3] != INFO_COLS:
        print(f"เตือน: หัวตาราง 3 คอลัมน์แรกควรเป็น {INFO_COLS} แต่พบ {header[:3]}")

    subject_names = [h for h in header[3:] if h]
    full_scores_row = ws[2]
    full_scores = {}
    for idx, name in enumerate(subject_names, start=4):  # column D=4 onward
        val = full_scores_row[idx - 1].value
        try:
            full_scores[name] = float(val) if val is not None else 100.0
        except (TypeError, ValueError):
            full_scores[name] = 100.0

    init_db()
    conn = get_conn()
    cur = conn.cursor()

    if replace:
        cur.execute("DELETE FROM scores")
        cur.execute("DELETE FROM students")
        cur.execute("DELETE FROM subjects")

    # upsert subjects (just registers the subject name; the max score for
    # this particular import is stored per score entry below, so two classes
    # using the same subject name with different max scores never clash)
    subject_ids = {}
    for name in subject_names:
        cur.execute(
            "INSERT INTO subjects(name, full_score) VALUES (?, ?) "
            "ON CONFLICT(name) DO NOTHING",
            (name, full_scores[name]),
        )
        cur.execute("SELECT id FROM subjects WHERE name=?", (name,))
        subject_ids[name] = cur.fetchone()["id"]

    imported = 0
    skipped = 0
    # data starts row 3; skip rows missing class/number/name
    for row in ws.iter_rows(min_row=3, values_only=False):
        class_name = row[0].value
        number = row[1].value
        full_name = row[2].value
        if not class_name or number is None or not full_name:
            skipped += 1
            continue
        class_name = str(class_name).strip()
        full_name = str(full_name).strip()
        try:
            number = int(number)
        except (TypeError, ValueError):
            skipped += 1
            continue

        cur.execute(
            "INSERT INTO students(class_name, number, full_name) VALUES (?, ?, ?) "
            "ON CONFLICT(class_name, number) DO UPDATE SET full_name=excluded.full_name",
            (class_name, number, full_name),
        )
        cur.execute(
            "SELECT id FROM students WHERE class_name=? AND number=?",
            (class_name, number),
        )
        student_id = cur.fetchone()["id"]

        for idx, name in enumerate(subject_names, start=4):
            cell = row[idx - 1]
            score = cell.value
            if score is None or score == "":
                continue
            try:
                score = float(score)
            except (TypeError, ValueError):
                continue
            cur.execute(
                "INSERT INTO scores(student_id, subject_id, score, full_score) VALUES (?, ?, ?, ?) "
                "ON CONFLICT(student_id, subject_id) DO UPDATE SET "
                "score=excluded.score, full_score=excluded.full_score",
                (student_id, subject_ids[name], score, full_scores[name]),
            )
        imported += 1

    conn.commit()
    conn.close()
    print(f"นำเข้าสำเร็จ: {imported} แถว, ข้าม: {skipped} แถว, วิชาทั้งหมด: {len(subject_names)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="นำเข้าคะแนนนักเรียนจาก Excel เข้าสู่ระบบ")
    parser.add_argument("file", help="พาธไปยังไฟล์ .xlsx ที่กรอกคะแนนแล้ว")
    parser.add_argument("--sheet", default=None, help="ชื่อชีทที่ต้องการอ่าน (ค่าเริ่มต้น: 'คะแนน')")
    parser.add_argument("--replace", action="store_true", help="ล้างข้อมูลเดิมทั้งหมดก่อนนำเข้าใหม่")
    args = parser.parse_args()

    try:
        import_scores(args.file, args.sheet, args.replace)
    except FileNotFoundError:
        print(f"ไม่พบไฟล์: {args.file}")
        sys.exit(1)
